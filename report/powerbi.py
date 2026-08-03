import os
import pandas as pd
from datetime import datetime


def export_results_to_excel(results: dict, filename: str = "analysis_export.xlsx") -> str:
    """Export analysis results to Excel for Power BI or local use."""
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Existing single-period sales growth (week) for backward compatibility
        growth_df = pd.DataFrame([results.get("growth", {})])
        # convert growth_pct from percent like 12.3 -> 0.123 for Excel percentage format
        if "growth_pct" in growth_df.columns:
            growth_df["growth_pct"] = growth_df["growth_pct"] / 100.0
        growth_df.to_excel(writer, sheet_name='Sales Growth', index=False)

        # Multi-period growth (week, month, quarter) if available
        growth_multi = results.get("growth_multi")
        if growth_multi:
            rows = []
            for period in ["week", "month", "quarter"]:
                entry = growth_multi.get(period, {"current": 0, "previous": 0, "growth_pct": 0})
                rows.append({
                    "Period": period.capitalize(),
                    "Current": entry.get("current", 0),
                    "Previous": entry.get("previous", 0),
                    # store as fractional value for Excel percent format
                    "GrowthPct": entry.get("growth_pct", 0) / 100.0,
                })
            gm_df = pd.DataFrame(rows)
            gm_df.to_excel(writer, sheet_name='Growth by Period', index=False)

        # Time series: weekly is expected in results
        if "weekly_sales" in results and getattr(results["weekly_sales"], "shape", (0,))[0] > 0:
            ws_df = results["weekly_sales"].copy()
            # ensure InvoiceDate is datetime
            ws_df["InvoiceDate"] = pd.to_datetime(ws_df["InvoiceDate"], dayfirst=True, errors="coerce")
            # add explicit week label (ISO week start date), month and quarter labels
            ws_df["WeekStart"] = ws_df["InvoiceDate"].dt.to_period('W').apply(lambda p: p.start_time.date())
            ws_df["Month"] = ws_df["InvoiceDate"].dt.to_period('M').apply(lambda p: p.start_time.strftime("%Y-%m"))
            def _quarter_label(ts):
                p = ts.to_period('Q')
                q = p.quarter
                return f"{p.start_time.year}Q{q}"

            ws_df["Quarter"] = ws_df["InvoiceDate"].apply(_quarter_label)

            # reorder columns for clarity
            cols = ["InvoiceDate", "WeekStart", "Month", "Quarter", "WeeklyRevenue"]
            for c in cols:
                if c not in ws_df.columns:
                    cols.remove(c)
            ws_df = ws_df[[c for c in cols if c in ws_df.columns]]
            ws_df.to_excel(writer, sheet_name='Weekly Revenue', index=False)

            # derive monthly from weekly by grouping by Month
            monthly = (
                ws_df.groupby('Month')["WeeklyRevenue"].sum()
                .reset_index()
                .rename(columns={"WeeklyRevenue": "MonthlyRevenue", "Month": "Period"})
            )
            monthly.to_excel(writer, sheet_name='Monthly Revenue', index=False)

            # derive quarterly from the Quarter label
            quarterly = (
                ws_df.groupby('Quarter')["WeeklyRevenue"].sum()
                .reset_index()
                .rename(columns={"WeeklyRevenue": "QuarterlyRevenue", "Quarter": "Period"})
            )
            quarterly.to_excel(writer, sheet_name='Quarterly Revenue', index=False)

        results["top_products"].to_excel(writer, sheet_name='Top Products', index=False)
        results["countries"].head(10).to_excel(writer, sheet_name='Countries', index=False)
        results["churn"].head(10).to_excel(writer, sheet_name='At-Risk Customers', index=False)
        results["anomalies"].to_excel(writer, sheet_name='Anomalies', index=False)

        # Apply number formats using openpyxl
        try:
            from openpyxl.utils import get_column_letter

            wb = writer.book

            # Format Sales Growth sheet growth_pct column as percentage
            if "Sales Growth" in writer.sheets and not growth_df.empty:
                ws = writer.sheets["Sales Growth"]
                if "growth_pct" in growth_df.columns:
                    col_idx = list(growth_df.columns).index("growth_pct") + 1
                    col_letter = get_column_letter(col_idx)
                    for row_idx in range(2, 2 + len(growth_df)):
                        cell = ws[f"{col_letter}{row_idx}"]
                        cell.number_format = '0.0%'

            # Format Growth by Period sheet: GrowthPct -> percentage, Current/Previous -> number with 2 decimals
            if growth_multi and "Growth by Period" in writer.sheets:
                ws2 = writer.sheets["Growth by Period"]
                cols = list(gm_df.columns)
                if "GrowthPct" in cols:
                    gcol = get_column_letter(cols.index("GrowthPct") + 1)
                    for r in range(2, 2 + len(gm_df)):
                        ws2[f"{gcol}{r}"].number_format = '0.0%'
                for name in ("Current", "Previous"):
                    if name in cols:
                        ccol = get_column_letter(cols.index(name) + 1)
                        for r in range(2, 2 + len(gm_df)):
                            ws2[f"{ccol}{r}"].number_format = '#,##0.00'
        except Exception:
            # If openpyxl formatting fails, we still return the file; formatting is optional
            pass

    print(f"✓ Data exported to {filename}")
    return filename


class PowerBIClient:
    """
    Handles Power BI integration for uploading data and sharing reports.
    
    Required environment variables:
      POWERBI_CLIENT_ID       - Azure AD App ID
      POWERBI_CLIENT_SECRET   - Azure AD App Password
      POWERBI_TENANT_ID       - Azure AD Tenant ID
      POWERBI_WORKSPACE_ID    - Power BI Workspace ID
      POWERBI_DATASET_ID      - Power BI Dataset ID (created in Power BI)
      POWERBI_REPORT_ID       - Power BI Report ID (created in Power BI)
    """
    
    BASE_URL = "https://api.powerbi.com/v1.0/myorg"
    TOKEN_URL = "https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"
    
    def __init__(self):
        self.client_id = os.environ.get("POWERBI_CLIENT_ID")
        self.client_secret = os.environ.get("POWERBI_CLIENT_SECRET")
        self.tenant_id = os.environ.get("POWERBI_TENANT_ID")
        self.workspace_id = os.environ.get("POWERBI_WORKSPACE_ID")
        self.dataset_id = os.environ.get("POWERBI_DATASET_ID")
        self.report_id = os.environ.get("POWERBI_REPORT_ID")
        self.access_token = None
        
        if not all([self.client_id, self.client_secret, self.tenant_id, 
                    self.workspace_id, self.dataset_id, self.report_id]):
            raise ValueError("Missing required Power BI environment variables")
    
    def authenticate(self):
        """Get access token from Azure AD."""
        try:
            import requests
        except ImportError as exc:
            raise ImportError(
                "Power BI integration requires the 'requests' package. "
                "Install it with: pip install requests"
            ) from exc

        url = self.TOKEN_URL.format(tenant=self.tenant_id)
        payload = {
            "client_id": self.client_id,
            "client_secret": self.client_secret,
            "scope": "https://analysis.windows.net/.default",
            "grant_type": "client_credentials"
        }
        
        response = requests.post(url, data=payload)
        if response.status_code != 200:
            raise Exception(f"Authentication failed: {response.text}")
        
        self.access_token = response.json()["access_token"]
        return self.access_token
    
    def _make_request(self, method, endpoint, data=None):
        """Make authenticated request to Power BI API."""
        if not self.access_token:
            self.authenticate()

        try:
            import requests
        except ImportError as exc:
            raise ImportError(
                "Power BI integration requires the 'requests' package. "
                "Install it with: pip install requests"
            ) from exc
        
        headers = {
            "Authorization": f"Bearer {self.access_token}",
            "Content-Type": "application/json"
        }
        
        url = f"{self.BASE_URL}{endpoint}"
        
        if method == "GET":
            response = requests.get(url, headers=headers)
        elif method == "POST":
            response = requests.post(url, headers=headers, json=data)
        elif method == "PUT":
            response = requests.put(url, headers=headers, json=data)
        else:
            raise ValueError(f"Unsupported method: {method}")
        
        return response
    
    def push_data(self, results: dict):
        """Push analysis results to Power BI dataset."""
        # Prepare data for pushing to Power BI
        payload = {
            "rows": [
                {
                    "Date": datetime.today().isoformat(),
                    "WeeklyRevenue": results["growth"]["current"],
                    "PreviousWeekRevenue": results["growth"]["previous"],
                    "GrowthPercentage": results["growth"]["growth_pct"],
                    "TopProductCount": len(results["top_products"]),
                    "CountriesCount": len(results["countries"]),
                    "AtRiskCustomers": len(results["churn"]),
                    "AnomaliesDetected": len(results["anomalies"]),
                }
            ]
        }
        
        endpoint = f"/groups/{self.workspace_id}/datasets/{self.dataset_id}/rows"
        response = self._make_request("POST", endpoint, payload)
        
        if response.status_code not in [200, 201]:
            raise Exception(f"Failed to push data: {response.text}")
        
        print("✓ Data pushed to Power BI dataset")
    
    def refresh_dataset(self):
        """Trigger a dataset refresh."""
        endpoint = f"/groups/{self.workspace_id}/datasets/{self.dataset_id}/refreshes"
        response = self._make_request("POST", endpoint, {})
        
        if response.status_code not in [200, 202]:
            raise Exception(f"Failed to refresh dataset: {response.text}")
        
        print("✓ Power BI dataset refresh initiated")
    
    def get_report_url(self):
        """Get the report embed URL."""
        endpoint = f"/groups/{self.workspace_id}/reports/{self.report_id}"
        response = self._make_request("GET", endpoint)
        
        if response.status_code != 200:
            raise Exception(f"Failed to get report URL: {response.text}")
        
        return response.json()["webUrl"]
    
    def export_to_excel(self, results: dict, filename: str = "analysis_export.xlsx"):
        """Export analysis results to Excel for Power BI to read."""
        return export_results_to_excel(results, filename)


def generate_powerbi_report(results: dict, recipient: str) -> tuple:
    """
    Generate Power BI report link and prepare for emailing.
    
    Returns:
        tuple: (report_url, export_filename)
    """
    try:
        client = PowerBIClient()
        
        # Export data to Excel
        export_file = client.export_to_excel(results)
        
        # Push data to Power BI
        client.push_data(results)
        
        # Refresh the dataset
        client.refresh_dataset()
        
        # Get report URL
        report_url = client.get_report_url()
        
        return report_url, export_file
    
    except Exception as e:
        print(f"ERROR: Power BI generation failed: {e}")
        print("Falling back to generating Excel export only...")
        
        export_file = export_results_to_excel(results, f"analysis_{datetime.today().strftime('%Y%m%d')}.xlsx")
        return None, export_file
